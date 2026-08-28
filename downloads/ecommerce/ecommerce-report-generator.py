"""
E-Ticaret Satış ve Görselleştirme Projesi
-----------------------------------------
1500 satırlık gerçekçi e-ticaret verisi üretir, pandas ile analiz eder ve
openpyxl ile kurumsal tasarımlı ECommerce_Sales_Report.xlsx dosyasını oluşturur.

Google Colab: Dosya üretildikten sonra indirme otomatik başlar.
Jupyter Notebook: Otomatik indirme tetiklenir ve ayrıca güvenli bir dosya linki gösterilir.
"""

from __future__ import annotations

import base64
import math
import os
import random
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# -----------------------------------------------------------------------------
# 1) PROJE AYARLARI
# -----------------------------------------------------------------------------
OUTPUT_FILE = Path("ECommerce_Sales_Report.xlsx")
ROW_COUNT = 1500
RANDOM_SEED = 42

# Kurumsal renk paleti: Gece mavisi + beyaz + turkuaz vurgu
NAVY = "0B1F3A"
NAVY_2 = "123B5D"
TEAL = "00A6A6"
GOLD = "F4B942"
WHITE = "FFFFFF"
LIGHT_BLUE = "DCEAF5"
LIGHT_GRAY = "F4F7FA"
MID_GRAY = "D0D7DE"
DARK_TEXT = "1F2937"
GREEN = "16A085"
RED = "C0392B"

CURRENCY_FORMAT = '₺ #,##0.00;[Red]-₺ #,##0.00'
CURRENCY_FORMAT_0 = '₺ #,##0;[Red]-₺ #,##0'
INTEGER_FORMAT = '#,##0'
PERCENT_FORMAT = '0.0%'
DATE_FORMAT = 'dd.mm.yyyy'

thin_gray = Side(style="thin", color=MID_GRAY)
medium_teal = Side(style="medium", color=TEAL)


# -----------------------------------------------------------------------------
# 2) GERÇEKÇİ E-TİCARET VERİSİ ÜRETİMİ
# -----------------------------------------------------------------------------
def simulate_ecommerce_data(n_rows: int = ROW_COUNT, seed: int = RANDOM_SEED) -> pd.DataFrame:
    """Mevsimsellik, şehir ve kategori ağırlıkları içeren sentetik veri üretir."""
    rng = random.Random(seed)

    categories = ["Teknoloji", "Moda", "Ev/Yaşam", "Kozmetik"]
    category_weights = [0.31, 0.29, 0.24, 0.16]

    cities = [
        "İstanbul", "Ankara", "İzmir", "Bursa", "Antalya", "Kocaeli",
        "Adana", "Konya", "Gaziantep", "Mersin", "Kayseri"
    ]
    city_weights = [0.30, 0.14, 0.12, 0.08, 0.08, 0.06, 0.06, 0.05, 0.04, 0.04, 0.03]

    # Kasım-Aralık kampanyaları ve hafta sonu alışveriş yoğunluğu modellenir.
    month_factor = {
        1: 0.84, 2: 0.88, 3: 0.95, 4: 0.98, 5: 1.02, 6: 1.05,
        7: 1.00, 8: 1.03, 9: 1.08, 10: 1.14, 11: 1.42, 12: 1.58,
    }
    weekday_factor = {0: 0.92, 1: 0.94, 2: 0.96, 3: 1.00, 4: 1.08, 5: 1.20, 6: 1.16}

    all_dates = pd.date_range("2025-01-01", "2025-12-31", freq="D")
    date_weights = [month_factor[d.month] * weekday_factor[d.weekday()] for d in all_dates]
    selected_dates = rng.choices(list(all_dates), weights=date_weights, k=n_rows)
    selected_dates.sort()

    price_profiles = {
        "Teknoloji": (900, 18_000, 4_800),
        "Moda": (180, 2_600, 720),
        "Ev/Yaşam": (300, 8_500, 1_850),
        "Kozmetik": (90, 1_900, 480),
    }
    quantity_profiles = {
        "Teknoloji": ([1, 2, 3], [0.72, 0.22, 0.06]),
        "Moda": ([1, 2, 3, 4, 5], [0.29, 0.32, 0.22, 0.12, 0.05]),
        "Ev/Yaşam": ([1, 2, 3, 4], [0.50, 0.30, 0.15, 0.05]),
        "Kozmetik": ([1, 2, 3, 4, 5, 6], [0.21, 0.28, 0.23, 0.15, 0.08, 0.05]),
    }

    records = []
    for index, order_date in enumerate(selected_dates, start=1):
        category = rng.choices(categories, weights=category_weights, k=1)[0]
        city = rng.choices(cities, weights=city_weights, k=1)[0]

        quantity_values, quantity_weights = quantity_profiles[category]
        quantity = rng.choices(quantity_values, weights=quantity_weights, k=1)[0]

        low_price, high_price, mode_price = price_profiles[category]
        unit_price = rng.triangular(low_price, high_price, mode_price)

        # Siparişlerin yaklaşık %43'ünde kampanya indirimi vardır.
        discount = rng.choices(
            [0.00, 0.05, 0.10, 0.15, 0.20],
            weights=[0.57, 0.12, 0.15, 0.10, 0.06],
            k=1,
        )[0]
        sales_amount = round(quantity * unit_price * (1 - discount), 2)

        records.append(
            {
                "Sipariş ID": f"SIP-{order_date:%Y%m}-{index:05d}",
                "Tarih": order_date,
                "Müşteri ID": f"MUS-{rng.randint(1, 620):04d}",
                "Ürün Kategorisi": category,
                "Satış Tutarı": sales_amount,
                "Miktar": int(quantity),
                "Şehir": city,
            }
        )

    df = pd.DataFrame(records)
    df["Tarih"] = pd.to_datetime(df["Tarih"])
    df["Satış Tutarı"] = df["Satış Tutarı"].astype(float).round(2)
    df["Miktar"] = df["Miktar"].astype(int)

    # Veri kalitesi kontrolleri: portföy projesinde veri güvenilirliğini gösterir.
    assert len(df) == n_rows, "Satır sayısı beklenen değerde değil."
    assert df["Sipariş ID"].is_unique, "Sipariş ID alanında tekrar var."
    assert not df.isna().any().any(), "Veri setinde eksik değer var."
    assert (df["Satış Tutarı"] > 0).all(), "Satış tutarı pozitif olmalıdır."
    assert (df["Miktar"] > 0).all(), "Miktar pozitif olmalıdır."

    return df


# -----------------------------------------------------------------------------
# 3) PANDAS İLE ANALİZ
# -----------------------------------------------------------------------------
def analyze_data(df: pd.DataFrame):
    """Kategori, aylık trend ve şehir analizlerini pandas ile hesaplar."""
    category_analysis = (
        df.groupby("Ürün Kategorisi", as_index=False)
        .agg(
            **{
                "Toplam Satış": ("Satış Tutarı", "sum"),
                "Toplam Miktar": ("Miktar", "sum"),
                "Sipariş Sayısı": ("Sipariş ID", "nunique"),
            }
        )
        .sort_values("Toplam Satış", ascending=False)
        .reset_index(drop=True)
    )
    category_analysis["Toplam Satış"] = category_analysis["Toplam Satış"].round(2)

    monthly_source = df.assign(Ay=df["Tarih"].dt.to_period("M").dt.to_timestamp())
    monthly_analysis = (
        monthly_source.groupby("Ay", as_index=False)
        .agg(
            **{
                "Aylık Satış": ("Satış Tutarı", "sum"),
                "Sipariş Sayısı": ("Sipariş ID", "nunique"),
            }
        )
        .sort_values("Ay")
        .reset_index(drop=True)
    )
    monthly_analysis["Aylık Satış"] = monthly_analysis["Aylık Satış"].round(2)

    month_names_tr = {
        1: "Oca", 2: "Şub", 3: "Mar", 4: "Nis", 5: "May", 6: "Haz",
        7: "Tem", 8: "Ağu", 9: "Eyl", 10: "Eki", 11: "Kas", 12: "Ara",
    }
    monthly_analysis["Ay Etiketi"] = monthly_analysis["Ay"].apply(
        lambda x: f"{month_names_tr[x.month]} {x.year}"
    )

    city_analysis = (
        df.groupby("Şehir", as_index=False)
        .agg(
            **{
                "Toplam Satış": ("Satış Tutarı", "sum"),
                "Sipariş Sayısı": ("Sipariş ID", "nunique"),
                "Toplam Miktar": ("Miktar", "sum"),
            }
        )
        .sort_values("Toplam Satış", ascending=False)
        .head(5)
        .reset_index(drop=True)
    )
    city_analysis["Toplam Satış"] = city_analysis["Toplam Satış"].round(2)

    return category_analysis, monthly_analysis, city_analysis


# -----------------------------------------------------------------------------
# 4) EXCEL TASARIM YARDIMCILARI
# -----------------------------------------------------------------------------
def fill_range(ws, cell_range: str, color: str) -> None:
    """Belirtilen hücre aralığına dolgu uygular."""
    fill = PatternFill("solid", fgColor=color)
    for row in ws[cell_range]:
        for cell in row:
            cell.fill = fill


def apply_title_band(ws, title: str, subtitle: str, end_column: int) -> None:
    """Her sayfada ortak kurumsal başlık bandını oluşturur."""
    end_letter = get_column_letter(end_column)
    ws.merge_cells(f"A1:{end_letter}2")
    ws["A1"] = title
    ws["A1"].font = Font(name="Aptos Display", size=20, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(f"A3:{end_letter}3")
    ws["A3"] = subtitle
    ws["A3"].font = Font(name="Aptos", size=10, italic=True, color=NAVY_2)
    ws["A3"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 10
    ws.row_dimensions[3].height = 22


def style_header_row(ws, row_number: int, start_col: int, end_col: int) -> None:
    """Tablo başlıklarını gece mavisi ve beyaz olarak biçimlendirir."""
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row_number, column=col)
        cell.fill = PatternFill("solid", fgColor=NAVY_2)
        cell.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=medium_teal)
    ws.row_dimensions[row_number].height = 28


def add_excel_table(ws, ref: str, name: str, style_name: str = "TableStyleMedium2") -> None:
    """Filtrelenebilir, bantlı Excel tablosu ekler."""
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name=style_name,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def auto_fit_columns(ws, min_width: int = 11, max_width: int = 32) -> None:
    """Birleştirilmiş başlıkları hariç tutarak kolon genişliklerini ayarlar."""
    for column_cells in ws.iter_cols():
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            if cell.coordinate in ws.merged_cells:
                continue
            value = cell.value
            if value is None or (isinstance(value, str) and value.startswith("=")):
                continue
            max_length = max(max_length, len(str(value)))
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, min_width), max_width)


def set_sheet_defaults(ws, zoom: int = 90) -> None:
    """Tüm sayfalarda ortak görünüm ve yazdırma ayarlarını uygular."""
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = zoom
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.outlinePr.summaryBelow = True


def style_card(ws, label_range: str, value_range: str, label: str, value, number_format=None) -> None:
    """Dashboard üzerinde iki parçalı KPI kartı oluşturur."""
    ws.merge_cells(label_range)
    ws.merge_cells(value_range)
    label_cell = ws[label_range.split(":")[0]]
    value_cell = ws[value_range.split(":")[0]]

    label_cell.value = label
    value_cell.value = value

    label_cell.fill = PatternFill("solid", fgColor=NAVY_2)
    label_cell.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
    label_cell.alignment = Alignment(horizontal="center", vertical="center")

    value_cell.fill = PatternFill("solid", fgColor=WHITE)
    value_cell.font = Font(name="Aptos Display", size=20, bold=True, color=NAVY)
    value_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if number_format:
        value_cell.number_format = number_format

    # Birleştirilmiş kartların dış hatlarını belirginleştirir.
    for cell_range in (label_range, value_range):
        min_col = ws[cell_range][0][0].column
        max_col = ws[cell_range][0][-1].column
        min_row = ws[cell_range][0][0].row
        max_row = ws[cell_range][-1][0].row
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = Border(
                    left=thin_gray if col == min_col else Side(style=None),
                    right=thin_gray if col == max_col else Side(style=None),
                    top=thin_gray if row == min_row else Side(style=None),
                    bottom=thin_gray if row == max_row else Side(style=None),
                )


def add_insight_box(ws, cell_range: str, text: str, fill_color: str = WHITE) -> None:
    """Dashboard için kısa yönetici bulgusu kutusu oluşturur."""
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = "• " + text
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.font = Font(name="Aptos", size=10, color=DARK_TEXT)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = Border(left=Side(style="medium", color=TEAL))


# -----------------------------------------------------------------------------
# 5) OPENPYXL İLE PROFESYONEL EXCEL RAPORU
# -----------------------------------------------------------------------------
def build_excel_report(
    df: pd.DataFrame,
    category_analysis: pd.DataFrame,
    monthly_analysis: pd.DataFrame,
    city_analysis: pd.DataFrame,
    output_file: Path = OUTPUT_FILE,
) -> Path:
    """Tüm analizleri, tabloları ve grafikleri içeren Excel raporunu üretir."""
    wb = Workbook()
    dashboard = wb.active
    dashboard.title = "Özet Rapor"
    raw_ws = wb.create_sheet("Ham Veri")
    category_ws = wb.create_sheet("Kategori Analizi")
    monthly_ws = wb.create_sheet("Aylık Trend")
    city_ws = wb.create_sheet("Şehir Analizi")

    # Dosya meta verileri portföy teslimini daha profesyonel gösterir.
    wb.properties.creator = "Yunus Emre Büyükgüler"
    wb.properties.title = "E-Ticaret Satış ve Görselleştirme Projesi"
    wb.properties.subject = "Python, pandas ve openpyxl ile satış analizi"
    wb.properties.description = "1500 satırlık sentetik e-ticaret verisi ve yönetici dashboardu."
    wb.properties.keywords = "Python, pandas, openpyxl, Excel, e-ticaret, dashboard, veri analizi"
    wb.properties.created = datetime.now()

    # Excel açıldığında formüllerin yeniden hesaplanmasını sağlar.
    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except AttributeError:
        pass

    # ----------------------------- HAM VERİ -----------------------------------
    apply_title_band(
        raw_ws,
        "HAM VERİ | E-TİCARET SİPARİŞLERİ",
        "1500 satırlık sentetik veri • Filtrelenebilir Excel tablosu • Kişisel veri içermez",
        7,
    )
    raw_header_row = 5
    raw_start_row = 6
    raw_ws.append([None] * 7)  # 4. satır için görsel boşluk
    raw_ws.append(list(df.columns))

    for record in df.itertuples(index=False, name=None):
        row = list(record)
        if isinstance(row[1], pd.Timestamp):
            row[1] = row[1].to_pydatetime()
        raw_ws.append(row)

    raw_end_row = raw_start_row + len(df) - 1
    style_header_row(raw_ws, raw_header_row, 1, 7)
    add_excel_table(raw_ws, f"A{raw_header_row}:G{raw_end_row}", "HamVeriTablosu")

    for row in range(raw_start_row, raw_end_row + 1):
        raw_ws.cell(row, 2).number_format = DATE_FORMAT
        raw_ws.cell(row, 5).number_format = CURRENCY_FORMAT
        raw_ws.cell(row, 6).number_format = INTEGER_FORMAT

    raw_ws.freeze_panes = f"A{raw_start_row}"
    raw_ws.auto_filter.ref = f"A{raw_header_row}:G{raw_end_row}"
    raw_ws.conditional_formatting.add(
        f"E{raw_start_row}:E{raw_end_row}",
        DataBarRule(start_type="min", end_type="max", color=TEAL, showValue=True),
    )
    raw_ws.conditional_formatting.add(
        f"F{raw_start_row}:F{raw_end_row}",
        ColorScaleRule(start_type="min", start_color=WHITE, end_type="max", end_color=LIGHT_BLUE),
    )
    raw_ws.column_dimensions["A"].width = 21
    raw_ws.column_dimensions["B"].width = 13
    raw_ws.column_dimensions["C"].width = 15
    raw_ws.column_dimensions["D"].width = 20
    raw_ws.column_dimensions["E"].width = 17
    raw_ws.column_dimensions["F"].width = 11
    raw_ws.column_dimensions["G"].width = 15
    set_sheet_defaults(raw_ws, zoom=85)

    # --------------------------- KATEGORİ ANALİZİ -----------------------------
    apply_title_band(
        category_ws,
        "KATEGORİ ANALİZİ",
        "Kategori bazında toplam satış, miktar, sipariş sayısı ve ciro payı",
        14,
    )
    category_headers = ["Ürün Kategorisi", "Toplam Satış", "Toplam Miktar", "Sipariş Sayısı", "Satış Payı"]
    for col, header in enumerate(category_headers, start=1):
        category_ws.cell(5, col, header)
    for i, record in enumerate(category_analysis.itertuples(index=False), start=6):
        category_ws.cell(i, 1, record[0])
        category_ws.cell(i, 2, float(record[1]))
        category_ws.cell(i, 3, int(record[2]))
        category_ws.cell(i, 4, int(record[3]))
        category_ws.cell(i, 5, f"=IFERROR(B{i}/SUM($B$6:$B$9),0)")

    style_header_row(category_ws, 5, 1, 5)
    add_excel_table(category_ws, "A5:E9", "KategoriAnaliziTablosu", "TableStyleMedium2")
    for row in range(6, 10):
        category_ws.cell(row, 2).number_format = CURRENCY_FORMAT
        category_ws.cell(row, 3).number_format = INTEGER_FORMAT
        category_ws.cell(row, 4).number_format = INTEGER_FORMAT
        category_ws.cell(row, 5).number_format = PERCENT_FORMAT

    pie = PieChart()
    pie.title = "Kategori Bazında Ciro Dağılımı"
    pie.style = 10
    pie.height = 8.0
    pie.width = 13.0
    pie.varyColors = True
    pie.legend = None
    pie.add_data(Reference(category_ws, min_col=2, min_row=5, max_row=9), titles_from_data=True)
    pie.set_categories(Reference(category_ws, min_col=1, min_row=6, max_row=9))
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showCatName = True
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = False
    pie.dataLabels.showSerName = False
    pie.dataLabels.showLegendKey = False
    pie.dataLabels.showLeaderLines = True
    category_ws.add_chart(pie, "G5")

    category_ws.conditional_formatting.add(
        "B6:B9", DataBarRule(start_type="min", end_type="max", color=TEAL, showValue=True)
    )
    category_ws.freeze_panes = "A6"
    auto_fit_columns(category_ws)
    category_ws.column_dimensions["A"].width = 21
    category_ws.column_dimensions["B"].width = 18
    set_sheet_defaults(category_ws)
    category_ws.print_area = "A1:N22"
    category_ws.page_setup.fitToHeight = 1

    # ----------------------------- AYLIK TREND --------------------------------
    apply_title_band(
        monthly_ws,
        "AYLIK SATIŞ TRENDİ",
        "2025 yılı aylık ciro, sipariş adedi, ortalama sepet ve büyüme görünümü",
        14,
    )
    monthly_headers = ["Ay", "Aylık Satış", "Sipariş Sayısı", "Ortalama Sepet", "Aylık Büyüme"]
    for col, header in enumerate(monthly_headers, start=1):
        monthly_ws.cell(5, col, header)

    for i, record in enumerate(monthly_analysis.itertuples(index=False), start=6):
        # record sırası: Ay, Aylık Satış, Sipariş Sayısı, Ay Etiketi
        monthly_ws.cell(i, 1, record[3])
        monthly_ws.cell(i, 2, float(record[1]))
        monthly_ws.cell(i, 3, int(record[2]))
        monthly_ws.cell(i, 4, f"=IFERROR(B{i}/C{i},0)")
        monthly_ws.cell(i, 5, "" if i == 6 else f'=IFERROR(B{i}/B{i-1}-1,"")')

    monthly_end_row = 5 + len(monthly_analysis)
    style_header_row(monthly_ws, 5, 1, 5)
    add_excel_table(monthly_ws, f"A5:E{monthly_end_row}", "AylikTrendTablosu", "TableStyleMedium2")
    for row in range(6, monthly_end_row + 1):
        monthly_ws.cell(row, 2).number_format = CURRENCY_FORMAT
        monthly_ws.cell(row, 3).number_format = INTEGER_FORMAT
        monthly_ws.cell(row, 4).number_format = CURRENCY_FORMAT
        monthly_ws.cell(row, 5).number_format = PERCENT_FORMAT

    line = LineChart()
    line.title = "Aylık Ciro Trendi (TL)"
    line.style = 13
    line.height = 8.0
    line.width = 14.5
    line.y_axis.title = "Ciro (TL)"
    line.x_axis.title = "Ay"
    line.y_axis.numFmt = CURRENCY_FORMAT_0
    line.legend = None
    line.add_data(
        Reference(monthly_ws, min_col=2, min_row=5, max_row=monthly_end_row),
        titles_from_data=True,
    )
    line.set_categories(Reference(monthly_ws, min_col=1, min_row=6, max_row=monthly_end_row))
    if line.series:
        series = line.series[0]
        series.graphicalProperties.line.solidFill = TEAL
        series.graphicalProperties.line.width = 28000
        series.marker.symbol = "circle"
        series.marker.size = 7
        series.marker.graphicalProperties.solidFill = GOLD
        series.marker.graphicalProperties.line.solidFill = NAVY
    monthly_ws.add_chart(line, "G5")

    monthly_ws.conditional_formatting.add(
        f"E7:E{monthly_end_row}",
        ColorScaleRule(
            start_type="min", start_color="F4CCCC",
            mid_type="percentile", mid_value=50, mid_color=WHITE,
            end_type="max", end_color="D9EAD3",
        ),
    )
    monthly_ws.freeze_panes = "A6"
    auto_fit_columns(monthly_ws)
    monthly_ws.column_dimensions["A"].width = 14
    monthly_ws.column_dimensions["B"].width = 18
    monthly_ws.column_dimensions["D"].width = 18
    set_sheet_defaults(monthly_ws)
    monthly_ws.print_area = "A1:N22"
    monthly_ws.page_setup.fitToHeight = 1

    # ----------------------------- ŞEHİR ANALİZİ ------------------------------
    apply_title_band(
        city_ws,
        "İLK 5 ŞEHİR ANALİZİ",
        "Ciroya göre lider şehirler, sipariş hacmi, miktar, ortalama sepet ve toplam pay",
        14,
    )
    city_headers = ["Şehir", "Toplam Satış", "Sipariş Sayısı", "Toplam Miktar", "Ortalama Sepet", "Ciro Payı"]
    for col, header in enumerate(city_headers, start=1):
        city_ws.cell(5, col, header)
    for i, record in enumerate(city_analysis.itertuples(index=False), start=6):
        city_ws.cell(i, 1, record[0])
        city_ws.cell(i, 2, float(record[1]))
        city_ws.cell(i, 3, int(record[2]))
        city_ws.cell(i, 4, int(record[3]))
        city_ws.cell(i, 5, f"=IFERROR(B{i}/C{i},0)")
        city_ws.cell(i, 6, f"=IFERROR(B{i}/SUM('Ham Veri'!$E$6:$E${raw_end_row}),0)")

    style_header_row(city_ws, 5, 1, 6)
    add_excel_table(city_ws, "A5:F10", "SehirAnaliziTablosu", "TableStyleMedium2")
    for row in range(6, 11):
        city_ws.cell(row, 2).number_format = CURRENCY_FORMAT
        city_ws.cell(row, 3).number_format = INTEGER_FORMAT
        city_ws.cell(row, 4).number_format = INTEGER_FORMAT
        city_ws.cell(row, 5).number_format = CURRENCY_FORMAT
        city_ws.cell(row, 6).number_format = PERCENT_FORMAT

    bar = BarChart()
    bar.type = "col"
    bar.grouping = "clustered"
    bar.overlap = 0
    bar.title = "En Çok Satış Yapılan İlk 5 Şehir"
    bar.style = 10
    bar.height = 8.0
    bar.width = 14.0
    bar.y_axis.title = "Ciro (TL)"
    bar.x_axis.title = "Şehir"
    bar.y_axis.numFmt = CURRENCY_FORMAT_0
    bar.legend = None
    bar.add_data(Reference(city_ws, min_col=2, min_row=5, max_row=10), titles_from_data=True)
    bar.set_categories(Reference(city_ws, min_col=1, min_row=6, max_row=10))
    if bar.series:
        bar.series[0].graphicalProperties.solidFill = TEAL
        bar.series[0].graphicalProperties.line.solidFill = NAVY
    city_ws.add_chart(bar, "H5")

    city_ws.conditional_formatting.add(
        "B6:B10", DataBarRule(start_type="min", end_type="max", color=TEAL, showValue=True)
    )
    city_ws.freeze_panes = "A6"
    auto_fit_columns(city_ws)
    city_ws.column_dimensions["A"].width = 16
    city_ws.column_dimensions["B"].width = 18
    city_ws.column_dimensions["E"].width = 18
    set_sheet_defaults(city_ws)
    city_ws.print_area = "A1:N22"
    city_ws.page_setup.fitToHeight = 1

    # ----------------------------- ÖZET RAPOR ---------------------------------
    fill_range(dashboard, "A1:L34", NAVY)
    dashboard.merge_cells("A1:L2")
    dashboard["A1"] = "E-TİCARET SATIŞ PERFORMANSI | YÖNETİCİ ÖZETİ"
    dashboard["A1"].font = Font(name="Aptos Display", size=22, bold=True, color=WHITE)
    dashboard["A1"].alignment = Alignment(horizontal="left", vertical="center")

    dashboard.merge_cells("A3:L3")
    dashboard["A3"] = "2025 • 1.500 Sipariş • Python (pandas) + Excel (openpyxl)"
    dashboard["A3"].font = Font(name="Aptos", size=10, italic=True, color=LIGHT_BLUE)
    dashboard["A3"].alignment = Alignment(horizontal="left", vertical="center")

    style_card(
        dashboard, "A5:C5", "A6:C8", "TOPLAM CİRO",
        f"=SUM('Ham Veri'!$E$6:$E${raw_end_row})", CURRENCY_FORMAT_0,
    )
    style_card(
        dashboard, "E5:G5", "E6:G8", "TOPLAM SİPARİŞ",
        f"=COUNTA('Ham Veri'!$A$6:$A${raw_end_row})", INTEGER_FORMAT,
    )
    style_card(
        dashboard, "I5:K5", "I6:K8", "TOPLAM ÜRÜN ADEDİ",
        f"=SUM('Ham Veri'!$F$6:$F${raw_end_row})", INTEGER_FORMAT,
    )
    style_card(
        dashboard, "A10:C10", "A11:C13", "ORTALAMA SEPET",
        "=IFERROR(A6/E6,0)", CURRENCY_FORMAT,
    )
    style_card(
        dashboard, "E10:G10", "E11:G13", "TEKİL MÜŞTERİ",
        int(df["Müşteri ID"].nunique()), INTEGER_FORMAT,
    )
    style_card(
        dashboard, "I10:K10", "I11:K13", "LİDER KATEGORİ",
        "='Kategori Analizi'!$A$6",
    )

    dashboard["A6"].comment = Comment("Formül kaynağı: 'Ham Veri' sayfasındaki Satış Tutarı alanı.", "Yunus Emre Büyükgüler")
    dashboard["E6"].comment = Comment("Formül kaynağı: benzersiz Sipariş ID kayıtları.", "Yunus Emre Büyükgüler")
    dashboard["A11"].comment = Comment("Toplam ciro / toplam sipariş.", "Yunus Emre Büyükgüler")

    # Dashboard üzerindeki aylık trend grafiği, Aylık Trend sayfasına bağlıdır.
    dash_line = LineChart()
    dash_line.title = "Aylık Ciro Görünümü"
    dash_line.style = 13
    dash_line.height = 7.2
    dash_line.width = 14.3
    dash_line.y_axis.numFmt = CURRENCY_FORMAT_0
    dash_line.legend = None
    dash_line.add_data(
        Reference(monthly_ws, min_col=2, min_row=5, max_row=monthly_end_row),
        titles_from_data=True,
    )
    dash_line.set_categories(Reference(monthly_ws, min_col=1, min_row=6, max_row=monthly_end_row))
    if dash_line.series:
        dash_series = dash_line.series[0]
        dash_series.graphicalProperties.line.solidFill = TEAL
        dash_series.graphicalProperties.line.width = 26000
        dash_series.marker.symbol = "circle"
        dash_series.marker.size = 6
        dash_series.marker.graphicalProperties.solidFill = GOLD
    dashboard.add_chart(dash_line, "A16")

    # Yöneticiye tek bakışta anlatılabilecek öne çıkan bulgular.
    total_sales = float(df["Satış Tutarı"].sum())
    top_category = category_analysis.iloc[0]
    top_month = monthly_analysis.loc[monthly_analysis["Aylık Satış"].idxmax()]
    top_city = city_analysis.iloc[0]
    top_five_share = float(city_analysis["Toplam Satış"].sum() / total_sales)
    monthly_growth = monthly_analysis["Aylık Satış"].pct_change()
    best_growth_index = monthly_growth.idxmax()
    best_growth_month = monthly_analysis.loc[best_growth_index, "Ay Etiketi"]
    best_growth_rate = float(monthly_growth.loc[best_growth_index])

    dashboard.merge_cells("I16:L17")
    dashboard["I16"] = "ÖNE ÇIKAN BULGULAR"
    dashboard["I16"].fill = PatternFill("solid", fgColor=TEAL)
    dashboard["I16"].font = Font(name="Aptos Display", size=13, bold=True, color=WHITE)
    dashboard["I16"].alignment = Alignment(horizontal="left", vertical="center")

    add_insight_box(
        dashboard, "I18:L20",
        f"Lider kategori {top_category['Ürün Kategorisi']}; toplam cironun %{top_category['Toplam Satış'] / total_sales * 100:.1f}'ini oluşturdu.",
    )
    add_insight_box(
        dashboard, "I21:L23",
        f"En güçlü ay {top_month['Ay Etiketi']}; aylık ciro ₺{top_month['Aylık Satış']:,.0f} seviyesine ulaştı.",
        LIGHT_BLUE,
    )
    add_insight_box(
        dashboard, "I24:L26",
        f"Lider şehir {top_city['Şehir']}; toplam satış ₺{top_city['Toplam Satış']:,.0f}.",
    )
    add_insight_box(
        dashboard, "I27:L29",
        f"İlk 5 şehir toplam cironun %{top_five_share * 100:.1f}'ini temsil ediyor.",
        LIGHT_BLUE,
    )
    add_insight_box(
        dashboard, "I30:L32",
        f"En yüksek aylık büyüme {best_growth_month} döneminde %{best_growth_rate * 100:.1f} oldu.",
    )

    dashboard.merge_cells("A33:L33")
    dashboard["A33"] = "Not: Veri Python ile simüle edilmiştir. Rapor, portföy ve eğitim amacıyla hazırlanmıştır."
    dashboard["A33"].font = Font(name="Aptos", size=9, italic=True, color=LIGHT_BLUE)
    dashboard["A33"].alignment = Alignment(horizontal="left", vertical="center")
    dashboard.row_dimensions[33].height = 18

    for col in range(1, 13):
        dashboard.column_dimensions[get_column_letter(col)].width = 12.5
    dashboard.column_dimensions["D"].width = 2.5
    dashboard.column_dimensions["H"].width = 2.5
    dashboard.column_dimensions["L"].width = 14
    dashboard.row_dimensions[1].height = 30
    dashboard.row_dimensions[2].height = 12
    dashboard.row_dimensions[3].height = 22
    for row in (6, 7, 8, 11, 12, 13):
        dashboard.row_dimensions[row].height = 22
    for row in range(18, 33):
        dashboard.row_dimensions[row].height = 18

    dashboard.freeze_panes = "A4"
    set_sheet_defaults(dashboard, zoom=85)
    dashboard.print_area = "A1:L33"
    dashboard.page_setup.fitToHeight = 1
    dashboard.page_margins.top = 0.25
    dashboard.page_margins.bottom = 0.25

    # Aktif sayfayı dashboard olarak belirler ve dosyayı kaydeder.
    wb.active = 0
    output_file = Path(output_file)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    return output_file.resolve()


# -----------------------------------------------------------------------------
# 6) DOSYA DOĞRULAMA VE NOTEBOOK İNDİRME
# -----------------------------------------------------------------------------
def validate_workbook(output_file: Path) -> None:
    """Sayfaları, satır sayısını, formülleri ve grafik nesnelerini kontrol eder."""
    wb = load_workbook(output_file, data_only=False)
    expected_sheets = ["Özet Rapor", "Ham Veri", "Kategori Analizi", "Aylık Trend", "Şehir Analizi"]
    assert wb.sheetnames == expected_sheets, "Sayfa sırası veya adları beklenen yapıda değil."
    assert wb["Ham Veri"].max_row == ROW_COUNT + 5, "Ham veri satır sayısı hatalı."
    assert len(wb["Kategori Analizi"]._charts) == 1, "Kategori pasta grafiği bulunamadı."
    assert len(wb["Aylık Trend"]._charts) == 1, "Aylık çizgi grafiği bulunamadı."
    assert len(wb["Şehir Analizi"]._charts) == 1, "Şehir sütun grafiği bulunamadı."
    assert len(wb["Özet Rapor"]._charts) == 1, "Dashboard trend grafiği bulunamadı."

    formula_errors = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and any(err in value for err in ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")):
                    formula_errors.append(f"{ws.title}!{cell.coordinate}: {value}")
    assert not formula_errors, "Formül hatası bulundu: " + "; ".join(formula_errors[:5])


def trigger_notebook_download(output_file: Path) -> None:
    """Colab'da indirir; Jupyter'da indirmeyi tetikler ve yedek link gösterir."""
    try:
        if "google.colab" in sys.modules or os.environ.get("COLAB_RELEASE_TAG"):
            from google.colab import files  # type: ignore

            files.download(str(output_file))
            return

        # Jupyter Notebook / JupyterLab ortamı
        get_ipython  # type: ignore[name-defined]  # noqa: F821
        from IPython.display import FileLink, Javascript, display

        encoded = base64.b64encode(output_file.read_bytes()).decode("ascii")
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        javascript = f"""
        (() => {{
            const link = document.createElement('a');
            link.href = 'data:{mime};base64,{encoded}';
            link.download = '{output_file.name}';
            document.body.appendChild(link);
            link.click();
            document.body.removeChild(link);
        }})();
        """
        display(Javascript(javascript))
        display(FileLink(str(output_file)))
    except (NameError, ImportError, OSError):
        print(f"Dosya hazır: {output_file}")


def main() -> None:
    """Uçtan uca veri üretimi, analiz, Excel raporlama ve indirme akışı."""
    df = simulate_ecommerce_data()
    category_analysis, monthly_analysis, city_analysis = analyze_data(df)
    output_path = build_excel_report(
        df,
        category_analysis,
        monthly_analysis,
        city_analysis,
        OUTPUT_FILE,
    )
    validate_workbook(output_path)

    total_sales = df["Satış Tutarı"].sum()
    print("✓ Rapor başarıyla oluşturuldu ve doğrulandı.")
    print(f"✓ Satır sayısı: {len(df):,}")
    print(f"✓ Toplam ciro: ₺{total_sales:,.2f}")
    print(f"✓ Dosya: {output_path.name}")
    trigger_notebook_download(output_path)


if __name__ == "__main__":
    main()
