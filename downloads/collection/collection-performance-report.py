"""
Tahsilat Performansı ve Çağrı Merkezi Verimlilik Analizi

Google Colab veya Jupyter Notebook üzerinde tek hücrede çalışacak şekilde
tasarlanmıştır. 2.500 satırlık sentetik Recall çağrı verisi üretir, KPI'ları
Pandas ile hesaplar ve kurumsal biçimlendirilmiş bir Excel raporu oluşturur.
"""

# -----------------------------------------------------------------------------
# 0) Gerekli paketleri kontrol et
# -----------------------------------------------------------------------------
import importlib.util
import os
import subprocess
import sys
import warnings

REQUIRED_PACKAGES = {
    "numpy": "numpy",
    "pandas": "pandas",
    "openpyxl": "openpyxl",
}

missing_packages = [
    pip_name
    for module_name, pip_name in REQUIRED_PACKAGES.items()
    if importlib.util.find_spec(module_name) is None
]
if missing_packages:
    subprocess.check_call(
        [sys.executable, "-m", "pip", "install", "-q", *missing_packages]
    )

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# -----------------------------------------------------------------------------
# 1) Ayarlar ve kurumsal renk paleti
# -----------------------------------------------------------------------------
RANDOM_SEED = 20260827
ROW_COUNT = 2500
DEBTOR_COUNT = 950
OUTPUT_FILE = "Collection_Performance_Report.xlsx"

rng = np.random.default_rng(RANDOM_SEED)

ANTHRACITE = "263238"
ANTHRACITE_2 = "37474F"
EMERALD = "009B77"
EMERALD_DARK = "007A5E"
MINT = "DDF6EE"
LIGHT_BG = "F4F7F6"
WHITE = "FFFFFF"
TEXT = "243238"
MID_GREY = "90A4AE"
LIGHT_GREY = "D9E2E5"
GOLD = "D6A84B"
RED = "C94C4C"

THIN_GREY = Side(style="thin", color=LIGHT_GREY)
MEDIUM_EMERALD = Side(style="medium", color=EMERALD)
CARD_BORDER = Border(
    left=MEDIUM_EMERALD,
    right=MEDIUM_EMERALD,
    top=MEDIUM_EMERALD,
    bottom=MEDIUM_EMERALD,
)

MONEY_FMT = '#,##0.00 "₺";[Red]-#,##0.00 "₺";-'
MONEY0_FMT = '#,##0 "₺";[Red]-#,##0 "₺";-'
PERCENT_FMT = "0.0%"
COUNT_FMT = "#,##0"
DECIMAL_FMT = "0.00"
DATE_FMT = "dd.mm.yyyy"

# -----------------------------------------------------------------------------
# 2) Gerçekçi ve ilişkili çağrı merkezi / tahsilat veri simülasyonu
# -----------------------------------------------------------------------------
representative_profiles = pd.DataFrame(
    [
        ("Ayşe Demir", "Ekip A", 0.82, 1.05),
        ("Burak Yılmaz", "Ekip A", 0.74, 1.00),
        ("Ceren Aydın", "Ekip A", 0.69, 0.96),
        ("Deniz Kaya", "Ekip A", 0.63, 1.02),
        ("Ece Şahin", "Ekip B", 0.79, 1.04),
        ("Furkan Arslan", "Ekip B", 0.71, 0.98),
        ("Gökçe Yıldız", "Ekip B", 0.66, 1.01),
        ("Hakan Koç", "Ekip B", 0.58, 0.95),
        ("İrem Çelik", "Ekip C", 0.76, 1.03),
        ("Kerem Öz", "Ekip C", 0.68, 0.99),
        ("Melis Aksoy", "Ekip C", 0.61, 1.00),
        ("Onur Kurt", "Ekip C", 0.55, 0.97),
    ],
    columns=["Temsilci_Adı", "Ekip", "Yetkinlik_Skoru", "Arama_Hacmi_Katsayısı"],
)

rep_names = representative_profiles["Temsilci_Adı"].to_numpy()
rep_volume_weights = representative_profiles["Arama_Hacmi_Katsayısı"].to_numpy()
rep_volume_weights = rep_volume_weights / rep_volume_weights.sum()
rep_profile_lookup = representative_profiles.set_index("Temsilci_Adı").to_dict("index")

debt_types = np.array(["Kredi Kartı", "İhtiyaç Kredisi", "Taşıt Kredisi"])
debtor_debt_types = rng.choice(debt_types, size=DEBTOR_COUNT, p=[0.46, 0.39, 0.15])

def simulate_balance(debt_type):
    if debt_type == "Kredi Kartı":
        return float(np.clip(rng.lognormal(np.log(22_000), 0.65), 2_500, 95_000))
    if debt_type == "İhtiyaç Kredisi":
        return float(np.clip(rng.lognormal(np.log(62_000), 0.60), 10_000, 260_000))
    return float(np.clip(rng.lognormal(np.log(155_000), 0.55), 30_000, 520_000))

initial_balances = np.array([simulate_balance(x) for x in debtor_debt_types])
remaining_balances = initial_balances.copy()
debtor_ids = np.array([f"BRC-{i:05d}" for i in range(1, DEBTOR_COUNT + 1)])

def debtor_segment(balance):
    if balance < 30_000:
        return "Düşük Bakiye"
    if balance < 100_000:
        return "Orta Bakiye"
    return "Yüksek Bakiye"

debtor_segments = np.array([debtor_segment(x) for x in initial_balances])
debtor_call_weights = np.sqrt(initial_balances)
debtor_call_weights = debtor_call_weights / debtor_call_weights.sum()

# Yönetim raporlarında eksik ay etkisini önlemek için son 12 tamamlanmış ay.
today = pd.Timestamp.today().normalize()
end_date = today - pd.offsets.MonthEnd(1)
start_date = (end_date - pd.DateOffset(months=11)).replace(day=1)
business_days = pd.bdate_range(start_date, end_date)
date_weights = np.linspace(0.85, 1.15, len(business_days))
date_weights = date_weights / date_weights.sum()
call_dates = np.sort(rng.choice(business_days.to_numpy(), size=ROW_COUNT, p=date_weights))

result_order = ["Tahsilat Yapıldı", "Söz Alındı", "Reddedildi", "Ulaşılamadı"]
type_adjustment = {
    "Kredi Kartı": np.array([0.025, 0.015, -0.010, -0.030]),
    "İhtiyaç Kredisi": np.array([0.000, 0.010, 0.005, -0.015]),
    "Taşıt Kredisi": np.array([-0.020, 0.025, 0.015, -0.020]),
}
segment_adjustment = {
    "Düşük Bakiye": np.array([0.025, 0.010, -0.010, -0.025]),
    "Orta Bakiye": np.array([0.000, 0.005, 0.005, -0.010]),
    "Yüksek Bakiye": np.array([-0.020, 0.015, 0.020, -0.015]),
}

records = []
for i, call_date in enumerate(call_dates, start=1):
    # Çok düşük bakiye kalmış bir borçlu seçilirse yeni borçlu seç.
    for _ in range(25):
        debtor_idx = int(rng.choice(DEBTOR_COUNT, p=debtor_call_weights))
        if remaining_balances[debtor_idx] >= 250:
            break

    debtor_id = debtor_ids[debtor_idx]
    debt_type = debtor_debt_types[debtor_idx]
    segment = debtor_segments[debtor_idx]
    balance_before_call = float(remaining_balances[debtor_idx])

    rep_name = str(rng.choice(rep_names, p=rep_volume_weights))
    profile = rep_profile_lookup[rep_name]
    team = profile["Ekip"]
    skill = float(profile["Yetkinlik_Skoru"])

    skill_delta = skill - 0.68
    probabilities = np.array([0.17, 0.25, 0.25, 0.33], dtype=float)
    probabilities += np.array(
        [0.16 * skill_delta, 0.08 * skill_delta, -0.07 * skill_delta, -0.17 * skill_delta]
    )
    probabilities += type_adjustment[debt_type]
    probabilities += segment_adjustment[segment]
    probabilities = np.clip(probabilities, 0.05, None)
    probabilities = probabilities / probabilities.sum()
    result = str(rng.choice(result_order, p=probabilities))

    duration_parameters = {
        "Ulaşılamadı": (1.15, 0.45),
        "Reddedildi": (3.40, 1.35),
        "Söz Alındı": (7.20, 2.10),
        "Tahsilat Yapıldı": (6.10, 1.85),
    }
    duration_mean, duration_std = duration_parameters[result]
    efficiency_factor = 1.08 - (0.15 * skill)
    call_duration = float(
        np.clip(rng.normal(duration_mean, duration_std) * efficiency_factor, 0.5, 15.0)
    )

    promise_amount = 0.0
    collection_amount = 0.0
    ptp_converted = 0

    if result == "Söz Alındı":
        promise_fraction = 0.03 + (0.44 * rng.beta(2.3, 7.2))
        promise_cap = {
            "Kredi Kartı": 35_000,
            "İhtiyaç Kredisi": 75_000,
            "Taşıt Kredisi": 135_000,
        }[debt_type]
        promise_amount = min(
            balance_before_call,
            promise_cap,
            max(250.0, balance_before_call * promise_fraction),
        )

        conversion_probability = 0.40 + (0.18 * (skill - 0.68))
        conversion_probability += {"Kredi Kartı": 0.035, "İhtiyaç Kredisi": 0.0, "Taşıt Kredisi": -0.03}[debt_type]
        conversion_probability += {"Düşük Bakiye": 0.03, "Orta Bakiye": 0.0, "Yüksek Bakiye": -0.025}[segment]
        conversion_probability = float(np.clip(conversion_probability, 0.25, 0.58))

        if rng.random() < conversion_probability:
            ptp_converted = 1
            collection_amount = min(
                balance_before_call,
                promise_amount * rng.uniform(0.72, 1.00),
            )

    elif result == "Tahsilat Yapıldı":
        payment_fraction = 0.04 + (0.34 * rng.beta(2.0, 6.2))
        payment_cap = {
            "Kredi Kartı": 40_000,
            "İhtiyaç Kredisi": 85_000,
            "Taşıt Kredisi": 150_000,
        }[debt_type]
        collection_amount = min(
            balance_before_call,
            payment_cap,
            max(250.0, balance_before_call * payment_fraction),
        )

    promise_amount = round(float(promise_amount), 2)
    collection_amount = round(float(collection_amount), 2)
    remaining_balances[debtor_idx] = max(0.0, balance_before_call - collection_amount)

    records.append(
        {
            "Çağrı_ID": f"CAG-{i:06d}",
            "Borçlu_ID": debtor_id,
            "Tarih": pd.Timestamp(call_date),
            "Temsilci_Adı": rep_name,
            "Ekip": team,
            "Çağrı_Süresi_Dk": round(call_duration, 2),
            "Arama_Sonucu": result,
            "Alınan_Söz_Tutarı": promise_amount,
            "Gerçekleşen_Tahsilat": collection_amount,
            "Borç_Tipi": debt_type,
            "Borçlu_Segmenti": segment,
            "Açık_Borç_Bakiyesi_Çağrı_Öncesi": round(balance_before_call, 2),
            "PTP_Dönüştü": ptp_converted,
            "Tahsilat_Başarı_Flag": int(collection_amount > 0),
            "Çağrı_Ayı": pd.Timestamp(call_date).strftime("%Y-%m"),
        }
    )

df = pd.DataFrame(records).sort_values(["Tarih", "Çağrı_ID"]).reset_index(drop=True)

# Mantıksal veri kontrolleri
assert len(df) == ROW_COUNT
assert df["Çağrı_ID"].is_unique
assert df["Çağrı_Süresi_Dk"].between(0.5, 15.0).all()
assert (df[["Alınan_Söz_Tutarı", "Gerçekleşen_Tahsilat"]] >= 0).all().all()
assert (df.loc[df["Arama_Sonucu"] != "Söz Alındı", "Alınan_Söz_Tutarı"] == 0).all()
assert (df.loc[df["Arama_Sonucu"].isin(["Reddedildi", "Ulaşılamadı"]), "Gerçekleşen_Tahsilat"] == 0).all()

# -----------------------------------------------------------------------------
# 3) Pandas ile KPI hesaplamaları
# -----------------------------------------------------------------------------
rep_group = df.groupby(["Temsilci_Adı", "Ekip"], as_index=False)
rep_perf = rep_group.agg(
    Toplam_Arama=("Çağrı_ID", "count"),
    Toplam_Söz_Sayısı=("Arama_Sonucu", lambda x: int((x == "Söz Alındı").sum())),
    Dönüşen_Söz_Sayısı=("PTP_Dönüştü", "sum"),
    Toplam_Söz_Tutarı=("Alınan_Söz_Tutarı", "sum"),
    Toplam_Tahsilat_Tutarı=("Gerçekleşen_Tahsilat", "sum"),
    Ortalama_Çağrı_Süresi_Dk=("Çağrı_Süresi_Dk", "mean"),
    Toplam_Çağrı_Dakikası=("Çağrı_Süresi_Dk", "sum"),
    Tahsilat_Başarı_Oranı=("Tahsilat_Başarı_Flag", "mean"),
)
rep_perf["Arama_Başına_Ort_Tahsilat"] = (
    rep_perf["Toplam_Tahsilat_Tutarı"] / rep_perf["Toplam_Arama"]
)
rep_perf["PTP_Dönüşüm_Oranı"] = (
    rep_perf["Dönüşen_Söz_Sayısı"]
    .div(rep_perf["Toplam_Söz_Sayısı"].replace(0, np.nan))
    .fillna(0)
)
rep_perf["Dakika_Başına_Tahsilat"] = (
    rep_perf["Toplam_Tahsilat_Tutarı"] / rep_perf["Toplam_Çağrı_Dakikası"]
)
rep_perf = rep_perf.drop(columns="Toplam_Çağrı_Dakikası")
rep_perf = rep_perf[
    [
        "Temsilci_Adı",
        "Ekip",
        "Toplam_Arama",
        "Toplam_Söz_Sayısı",
        "Dönüşen_Söz_Sayısı",
        "Toplam_Söz_Tutarı",
        "Toplam_Tahsilat_Tutarı",
        "Arama_Başına_Ort_Tahsilat",
        "Ortalama_Çağrı_Süresi_Dk",
        "PTP_Dönüşüm_Oranı",
        "Tahsilat_Başarı_Oranı",
        "Dakika_Başına_Tahsilat",
    ]
].sort_values("Toplam_Tahsilat_Tutarı", ascending=False).reset_index(drop=True)

def performance_summary(group_column):
    grouped = df.groupby(group_column, as_index=False).agg(
        Toplam_Arama=("Çağrı_ID", "count"),
        Toplam_Söz_Sayısı=("Arama_Sonucu", lambda x: int((x == "Söz Alındı").sum())),
        Dönüşen_Söz_Sayısı=("PTP_Dönüştü", "sum"),
        Toplam_Tahsilat_Tutarı=("Gerçekleşen_Tahsilat", "sum"),
        Ortalama_Çağrı_Süresi_Dk=("Çağrı_Süresi_Dk", "mean"),
        Toplam_Çağrı_Dakikası=("Çağrı_Süresi_Dk", "sum"),
        Tahsilat_Başarı_Oranı=("Tahsilat_Başarı_Flag", "mean"),
    )
    grouped["Arama_Başına_Ort_Tahsilat"] = grouped["Toplam_Tahsilat_Tutarı"] / grouped["Toplam_Arama"]
    grouped["PTP_Dönüşüm_Oranı"] = (
        grouped["Dönüşen_Söz_Sayısı"]
        .div(grouped["Toplam_Söz_Sayısı"].replace(0, np.nan))
        .fillna(0)
    )
    grouped["Dakika_Başına_Tahsilat"] = grouped["Toplam_Tahsilat_Tutarı"] / grouped["Toplam_Çağrı_Dakikası"]
    return grouped.drop(columns=["Toplam_Söz_Sayısı", "Dönüşen_Söz_Sayısı", "Toplam_Çağrı_Dakikası"])

team_perf = performance_summary("Ekip").sort_values("Toplam_Tahsilat_Tutarı", ascending=False).reset_index(drop=True)
debt_perf = performance_summary("Borç_Tipi").sort_values("Toplam_Tahsilat_Tutarı", ascending=False).reset_index(drop=True)
segment_perf = performance_summary("Borçlu_Segmenti").sort_values("Toplam_Tahsilat_Tutarı", ascending=False).reset_index(drop=True)

call_result_perf = (
    df.groupby("Arama_Sonucu", as_index=False)
    .agg(
        Arama_Sayısı=("Çağrı_ID", "count"),
        Ortalama_Çağrı_Süresi_Dk=("Çağrı_Süresi_Dk", "mean"),
        Toplam_Tahsilat_Tutarı=("Gerçekleşen_Tahsilat", "sum"),
    )
    .set_index("Arama_Sonucu")
    .reindex(["Ulaşılamadı", "Söz Alındı", "Reddedildi", "Tahsilat Yapıldı"])
    .reset_index()
)
call_result_perf["Yüzde"] = call_result_perf["Arama_Sayısı"] / ROW_COUNT
call_result_perf = call_result_perf[
    ["Arama_Sonucu", "Arama_Sayısı", "Yüzde", "Ortalama_Çağrı_Süresi_Dk", "Toplam_Tahsilat_Tutarı"]
]

monthly_perf = (
    df.groupby("Çağrı_Ayı", as_index=False)
    .agg(
        Toplam_Tahsilat_Tutarı=("Gerçekleşen_Tahsilat", "sum"),
        Toplam_Arama=("Çağrı_ID", "count"),
    )
    .sort_values("Çağrı_Ayı")
)

total_calls = int(len(df))
total_collection = float(df["Gerçekleşen_Tahsilat"].sum())
total_promises = int((df["Arama_Sonucu"] == "Söz Alındı").sum())
converted_promises = int(df["PTP_Dönüştü"].sum())
overall_ptp_rate = converted_promises / total_promises if total_promises else 0.0
average_call_duration = float(df["Çağrı_Süresi_Dk"].mean())
overall_success_rate = float(df["Tahsilat_Başarı_Flag"].mean())

# -----------------------------------------------------------------------------
# 4) Veri odaklı yönetici çıkarımları
# -----------------------------------------------------------------------------
def tr_money(value):
    return f"{value:,.0f}".replace(",", ".") + " ₺"

def tr_percent(value):
    return f"{value * 100:.1f}".replace(".", ",") + "%"

best_rep = rep_perf.iloc[0]
best_debt = debt_perf.sort_values(["Tahsilat_Başarı_Oranı", "Toplam_Tahsilat_Tutarı"], ascending=False).iloc[0]
best_team = team_perf.sort_values("Dakika_Başına_Tahsilat", ascending=False).iloc[0]
coaching_team = team_perf.sort_values("Dakika_Başına_Tahsilat", ascending=True).iloc[0]

insights = [
    (
        f"1. Dengeli prim modeli: {best_rep['Temsilci_Adı']} {tr_money(best_rep['Toplam_Tahsilat_Tutarı'])} ile liderdir. "
        f"Primin %50'si toplam tahsilat, %30'u PTP dönüşümü, %20'si dakika başına tahsilat üzerinden; görüşme kalitesi ve mevzuata uyum eşikleri korunarak hesaplanmalıdır."
    ),
    (
        f"2. Segment bazlı arama stratejisi: {best_debt['Borç_Tipi']} portföyü {tr_percent(best_debt['Tahsilat_Başarı_Oranı'])} tahsilat başarı oranıyla öne çıkmaktadır "
        f"(genel oran: {tr_percent(overall_success_rate)}). Yüksek performanslı temsilcilerin bu gruptaki başarılı konuşma akışları standartlaştırılmalı ve benzer dosyalara öncelikli kapasite ayrılmalıdır."
    ),
    (
        f"3. Koçluk ve kapasite: {best_team['Ekip']} dakika başına {tr_money(best_team['Dakika_Başına_Tahsilat'])} üretirken {coaching_team['Ekip']} için bu değer "
        f"{tr_money(coaching_team['Dakika_Başına_Tahsilat'])} seviyesindedir. Çağrı dinleme, itiraz karşılama ve kapanış koçluğu ile ekipler arası verimlilik farkı azaltılmalıdır."
    ),
]

# -----------------------------------------------------------------------------
# 5) openpyxl ile kurumsal Excel raporunun oluşturulması
# -----------------------------------------------------------------------------
wb = Workbook()
ws_dashboard = wb.active
ws_dashboard.title = "Özet Dashboard"
ws_rep = wb.create_sheet("Temsilci Performansı")
ws_call = wb.create_sheet("Çağrı Analizi")
ws_raw = wb.create_sheet("Ham Veri")

wb.properties.creator = "Yunus Emre Büyükgüler"
wb.properties.title = "Tahsilat Performansı ve Çağrı Merkezi Verimlilik Analizi"
wb.properties.subject = "NPL çağrı merkezi ve tahsilat KPI analizi"
wb.properties.description = "Sentetik Recall sistemi verisi ile hazırlanmış mülakat portföy çalışması."
try:
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
except Exception:
    pass

def py_value(value):
    if value is None:
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if isinstance(value, np.generic):
        return value.item()
    if pd.isna(value):
        return None
    return value

def style_sheet_base(ws, tab_color=EMERALD, zoom=90):
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = zoom
    ws.sheet_properties.tabColor = tab_color
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.oddHeader.center.text = "Tahsilat Performansı ve Çağrı Merkezi Verimlilik Analizi"
    ws.oddFooter.center.text = "Sayfa &P / &N"
    ws.oddFooter.right.text = "Portföy Çalışması"

def add_title(ws, title, subtitle, end_col):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    ws["A1"] = title
    ws["A1"].font = Font(name="Aptos Display", size=20, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=ANTHRACITE)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 34
    for col in range(1, end_col + 1):
        ws.cell(1, col).fill = PatternFill("solid", fgColor=ANTHRACITE)

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Aptos", size=11, color=WHITE)
    ws["A2"].fill = PatternFill("solid", fgColor=EMERALD_DARK)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 24
    for col in range(1, end_col + 1):
        ws.cell(2, col).fill = PatternFill("solid", fgColor=EMERALD_DARK)

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=end_col)
    ws["A3"] = (
        f"Recall sistemi senaryosu | {ROW_COUNT:,} çağrı | Analiz dönemi: "
        f"{start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')} | Sentetik veri"
    ).replace(",", ".")
    ws["A3"].font = Font(name="Aptos", size=9, italic=True, color=TEXT)
    ws["A3"].fill = PatternFill("solid", fgColor=MINT)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 21
    for col in range(1, end_col + 1):
        ws.cell(3, col).fill = PatternFill("solid", fgColor=MINT)

def add_section_header(ws, row, start_col, end_col, text):
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell = ws.cell(row, start_col, text)
    cell.font = Font(name="Aptos", size=11, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=ANTHRACITE_2)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 23
    for col in range(start_col, end_col + 1):
        ws.cell(row, col).fill = PatternFill("solid", fgColor=ANTHRACITE_2)

def add_kpi_card(ws, start_col, end_col, label, value, number_format):
    ws.merge_cells(start_row=5, start_column=start_col, end_row=6, end_column=end_col)
    ws.merge_cells(start_row=7, start_column=start_col, end_row=9, end_column=end_col)
    label_cell = ws.cell(5, start_col, label)
    value_cell = ws.cell(7, start_col, value)
    for row in range(5, 7):
        for col in range(start_col, end_col + 1):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=EMERALD)
            ws.cell(row, col).border = CARD_BORDER
    for row in range(7, 10):
        for col in range(start_col, end_col + 1):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=ANTHRACITE)
            ws.cell(row, col).border = CARD_BORDER
    label_cell.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
    label_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    value_cell.font = Font(name="Aptos Display", size=20, bold=True, color=WHITE)
    value_cell.alignment = Alignment(horizontal="center", vertical="center")
    value_cell.number_format = number_format

def write_dataframe_table(
    ws,
    data,
    start_row,
    start_col,
    table_name,
    money_columns=None,
    percent_columns=None,
    decimal_columns=None,
    integer_columns=None,
    date_columns=None,
    table_style="TableStyleMedium4",
):
    money_columns = set(money_columns or [])
    percent_columns = set(percent_columns or [])
    decimal_columns = set(decimal_columns or [])
    integer_columns = set(integer_columns or [])
    date_columns = set(date_columns or [])

    for offset, column_name in enumerate(data.columns):
        cell = ws.cell(start_row, start_col + offset, column_name)
        cell.font = Font(name="Aptos", size=9, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=EMERALD_DARK)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=MEDIUM_EMERALD)

    for row_offset, row_values in enumerate(data.itertuples(index=False, name=None), start=1):
        for col_offset, raw_value in enumerate(row_values):
            cell = ws.cell(start_row + row_offset, start_col + col_offset, py_value(raw_value))
            column_name = data.columns[col_offset]
            cell.font = Font(name="Aptos", size=9, color=TEXT)
            cell.alignment = Alignment(
                horizontal="left" if isinstance(cell.value, str) else "right",
                vertical="center",
            )
            if column_name in money_columns:
                cell.number_format = MONEY_FMT
            elif column_name in percent_columns:
                cell.number_format = PERCENT_FMT
            elif column_name in decimal_columns:
                cell.number_format = DECIMAL_FMT
            elif column_name in integer_columns:
                cell.number_format = COUNT_FMT
            elif column_name in date_columns:
                cell.number_format = DATE_FMT

    end_row = start_row + len(data)
    end_col = start_col + len(data.columns) - 1
    ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name=table_style,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    ws.row_dimensions[start_row].height = 33
    return end_row, end_col

def set_widths(ws, width_map):
    for column, width in width_map.items():
        ws.column_dimensions[column].width = width

def style_chart(chart, title, width=16, height=8, legend_position="b"):
    chart.title = title
    chart.style = 10
    chart.width = width
    chart.height = height
    chart.legend.position = legend_position
    chart.display_blanks = "gap"

# ----- Özet Dashboard ---------------------------------------------------------
style_sheet_base(ws_dashboard, tab_color=EMERALD, zoom=85)
add_title(
    ws_dashboard,
    "TAHSİLAT PERFORMANSI VE ÇAĞRI MERKEZİ VERİMLİLİK ANALİZİ",
    "Yönetim sorusu: Hangi ekip, hangi borçlu grubundan, ne kadar sürede tahsilat yapıyor?",
    16,
)
set_widths(ws_dashboard, {col: 13 for col in [get_column_letter(i) for i in range(1, 17)]})
ws_dashboard.column_dimensions["A"].width = 15
ws_dashboard.column_dimensions["B"].width = 22
ws_dashboard.column_dimensions["C"].width = 13
ws_dashboard.column_dimensions["D"].width = 19
ws_dashboard.column_dimensions["E"].width = 18
ws_dashboard.column_dimensions["F"].width = 15
ws_dashboard.column_dimensions["G"].width = 17
ws_dashboard.column_dimensions["H"].width = 18

add_kpi_card(ws_dashboard, 1, 4, "TOPLAM ARAMA", total_calls, COUNT_FMT)
add_kpi_card(ws_dashboard, 5, 8, "TOPLAM GERÇEKLEŞEN TAHSİLAT", total_collection, MONEY0_FMT)
add_kpi_card(ws_dashboard, 9, 12, "GENEL PTP DÖNÜŞÜM ORANI", overall_ptp_rate, PERCENT_FMT)
add_kpi_card(ws_dashboard, 13, 16, "ORTALAMA ÇAĞRI SÜRESİ", average_call_duration, '0.00 "dk"')
for row in range(5, 10):
    ws_dashboard.row_dimensions[row].height = 22

add_section_header(ws_dashboard, 11, 1, 8, "En Yüksek Tahsilat Yapan İlk 3 Temsilci")
add_section_header(ws_dashboard, 11, 9, 16, "Ekip Verimliliği")

top3 = rep_perf.head(3).copy()
top3.insert(0, "Sıra", range(1, len(top3) + 1))
top3 = top3[
    [
        "Sıra",
        "Temsilci_Adı",
        "Ekip",
        "Toplam_Tahsilat_Tutarı",
        "Arama_Başına_Ort_Tahsilat",
        "PTP_Dönüşüm_Oranı",
        "Ortalama_Çağrı_Süresi_Dk",
        "Dakika_Başına_Tahsilat",
    ]
]
write_dataframe_table(
    ws_dashboard,
    top3,
    12,
    1,
    "DashboardTop3Table",
    money_columns={"Toplam_Tahsilat_Tutarı", "Arama_Başına_Ort_Tahsilat", "Dakika_Başına_Tahsilat"},
    percent_columns={"PTP_Dönüşüm_Oranı"},
    decimal_columns={"Ortalama_Çağrı_Süresi_Dk"},
    integer_columns={"Sıra"},
)

team_dashboard = team_perf[
    [
        "Ekip",
        "Toplam_Arama",
        "Toplam_Tahsilat_Tutarı",
        "Arama_Başına_Ort_Tahsilat",
        "Ortalama_Çağrı_Süresi_Dk",
        "PTP_Dönüşüm_Oranı",
        "Tahsilat_Başarı_Oranı",
        "Dakika_Başına_Tahsilat",
    ]
]
write_dataframe_table(
    ws_dashboard,
    team_dashboard,
    12,
    9,
    "DashboardTeamTable",
    money_columns={"Toplam_Tahsilat_Tutarı", "Arama_Başına_Ort_Tahsilat", "Dakika_Başına_Tahsilat"},
    percent_columns={"PTP_Dönüşüm_Oranı", "Tahsilat_Başarı_Oranı"},
    decimal_columns={"Ortalama_Çağrı_Süresi_Dk"},
    integer_columns={"Toplam_Arama"},
)

add_section_header(ws_dashboard, 17, 1, 16, "Aylık Tahsilat Trendi")
monthly_end_row, _ = write_dataframe_table(
    ws_dashboard,
    monthly_perf,
    18,
    1,
    "MonthlyTrendTable",
    money_columns={"Toplam_Tahsilat_Tutarı"},
    integer_columns={"Toplam_Arama"},
)

monthly_chart = LineChart()
monthly_data = Reference(ws_dashboard, min_col=2, min_row=18, max_row=monthly_end_row)
monthly_categories = Reference(ws_dashboard, min_col=1, min_row=19, max_row=monthly_end_row)
monthly_chart.add_data(monthly_data, titles_from_data=True)
monthly_chart.set_categories(monthly_categories)
style_chart(monthly_chart, "Aylık Gerçekleşen Tahsilat Trendi (₺)", width=19, height=8.2, legend_position="b")
monthly_chart.y_axis.title = "Tahsilat (₺)"
monthly_chart.y_axis.numFmt = MONEY0_FMT
monthly_chart.x_axis.title = "Ay"
monthly_chart.series[0].graphicalProperties.line.solidFill = EMERALD
monthly_chart.series[0].graphicalProperties.line.width = 30000
monthly_chart.series[0].marker.symbol = "circle"
monthly_chart.series[0].marker.size = 6
monthly_chart.anchor = "E18"
ws_dashboard.add_chart(monthly_chart)

insight_header_row = max(34, monthly_end_row + 2)
add_section_header(ws_dashboard, insight_header_row, 1, 16, "Yönetici Çıkarımları ve Operasyonel Öneriler")
for index, insight in enumerate(insights):
    start_row_insight = insight_header_row + 1 + (index * 2)
    end_row_insight = start_row_insight + 1
    ws_dashboard.merge_cells(start_row=start_row_insight, start_column=1, end_row=end_row_insight, end_column=16)
    cell = ws_dashboard.cell(start_row_insight, 1, insight)
    cell.font = Font(name="Aptos", size=10, bold=(index == 0), color=TEXT)
    cell.fill = PatternFill("solid", fgColor=MINT if index % 2 == 0 else LIGHT_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = Border(left=MEDIUM_EMERALD, bottom=THIN_GREY)
    for row in range(start_row_insight, end_row_insight + 1):
        ws_dashboard.row_dimensions[row].height = 26
        for col in range(1, 17):
            ws_dashboard.cell(row, col).fill = PatternFill("solid", fgColor=MINT if index % 2 == 0 else LIGHT_BG)

note_row = insight_header_row + 8
ws_dashboard.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=16)
ws_dashboard.cell(note_row, 1, "Not: Tüm kişi, borçlu ve finansal değerler mülakat portföyü için sentetik olarak üretilmiştir; gerçek müşteri verisi içermez.")
ws_dashboard.cell(note_row, 1).font = Font(name="Aptos", size=8, italic=True, color=ANTHRACITE_2)
ws_dashboard.cell(note_row, 1).alignment = Alignment(horizontal="left")
ws_dashboard.freeze_panes = "A5"
ws_dashboard.auto_filter.ref = f"A18:C{monthly_end_row}"
ws_dashboard.print_area = f"A1:P{note_row}"

# ----- Temsilci Performansı --------------------------------------------------
style_sheet_base(ws_rep, tab_color=ANTHRACITE, zoom=82)
add_title(
    ws_rep,
    "TEMSİLCİ PERFORMANSI",
    "Personel bazında tahsilat, PTP dönüşümü, çağrı süresi ve dakika başına verimlilik karşılaştırması",
    28,
)
rep_end_row, rep_end_col = write_dataframe_table(
    ws_rep,
    rep_perf,
    5,
    1,
    "RepresentativePerformanceTable",
    money_columns={
        "Toplam_Söz_Tutarı",
        "Toplam_Tahsilat_Tutarı",
        "Arama_Başına_Ort_Tahsilat",
        "Dakika_Başına_Tahsilat",
    },
    percent_columns={"PTP_Dönüşüm_Oranı", "Tahsilat_Başarı_Oranı"},
    decimal_columns={"Ortalama_Çağrı_Süresi_Dk"},
    integer_columns={"Toplam_Arama", "Toplam_Söz_Sayısı", "Dönüşen_Söz_Sayısı"},
)
set_widths(
    ws_rep,
    {
        "A": 21,
        "B": 12,
        "C": 14,
        "D": 17,
        "E": 18,
        "F": 20,
        "G": 23,
        "H": 23,
        "I": 22,
        "J": 20,
        "K": 21,
        "L": 22,
        "M": 3,
    },
)
for col in range(14, 29):
    ws_rep.column_dimensions[get_column_letter(col)].width = 12

ws_rep.conditional_formatting.add(
    f"G6:G{rep_end_row}",
    DataBarRule(start_type="min", end_type="max", color=EMERALD, showValue=True),
)
ws_rep.conditional_formatting.add(
    f"J6:J{rep_end_row}",
    ColorScaleRule(
        start_type="min", start_color="FADBD8",
        mid_type="percentile", mid_value=50, mid_color="FFF2CC",
        end_type="max", end_color="C6EFCE",
    ),
)
for row in range(6, min(9, rep_end_row + 1)):
    for col in range(1, rep_end_col + 1):
        ws_rep.cell(row, col).font = Font(name="Aptos", size=9, bold=True, color=TEXT)

rep_chart = BarChart()
rep_chart.type = "col"
rep_chart.grouping = "clustered"
rep_chart.overlap = 0
rep_chart_data = Reference(ws_rep, min_col=6, max_col=7, min_row=5, max_row=rep_end_row)
rep_chart_categories = Reference(ws_rep, min_col=1, min_row=6, max_row=rep_end_row)
rep_chart.add_data(rep_chart_data, titles_from_data=True)
rep_chart.set_categories(rep_chart_categories)
style_chart(rep_chart, "Söz Tutarı ve Gerçekleşen Tahsilat Karşılaştırması (₺)", width=18.5, height=10.5, legend_position="b")
rep_chart.y_axis.title = "Tutar (₺)"
rep_chart.y_axis.numFmt = MONEY0_FMT
rep_chart.x_axis.title = "Müşteri Temsilcisi"
rep_chart.series[0].graphicalProperties.solidFill = MID_GREY
rep_chart.series[0].graphicalProperties.line.solidFill = MID_GREY
rep_chart.series[1].graphicalProperties.solidFill = EMERALD
rep_chart.series[1].graphicalProperties.line.solidFill = EMERALD_DARK
rep_chart.anchor = "N5"
ws_rep.add_chart(rep_chart)

definition_row = max(22, rep_end_row + 4)
add_section_header(ws_rep, definition_row, 1, 12, "KPI Tanımları")
definitions = [
    "PTP Dönüşüm Oranı = Tahsilata dönüşen söz adedi / Toplam söz adedi (adet bazlı).",
    "Arama Başına Ortalama Tahsilat = Toplam gerçekleşen tahsilat / Toplam arama sayısı.",
    "Dakika Başına Tahsilat = Toplam gerçekleşen tahsilat / Toplam çağrı dakikası.",
    "Tahsilat Başarı Oranı = Tahsilat üreten çağrı sayısı / Toplam çağrı sayısı.",
]
for idx, definition in enumerate(definitions, start=definition_row + 1):
    ws_rep.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=12)
    ws_rep.cell(idx, 1, f"• {definition}")
    ws_rep.cell(idx, 1).font = Font(name="Aptos", size=9, color=TEXT)
    ws_rep.cell(idx, 1).fill = PatternFill("solid", fgColor=LIGHT_BG)
    ws_rep.cell(idx, 1).alignment = Alignment(wrap_text=True, vertical="center")
    ws_rep.row_dimensions[idx].height = 22
ws_rep.freeze_panes = "C6"
ws_rep.print_area = f"A1:AB{definition_row + len(definitions)}"

# ----- Çağrı Analizi ---------------------------------------------------------
style_sheet_base(ws_call, tab_color=EMERALD_DARK, zoom=85)
add_title(
    ws_call,
    "ÇAĞRI ANALİZİ",
    "Arama sonucu, borç tipi ve borçlu segmenti bazında dağılım ve tahsilat başarısı",
    16,
)
set_widths(
    ws_call,
    {
        "A": 22,
        "B": 17,
        "C": 14,
        "D": 24,
        "E": 23,
        "F": 20,
        "G": 21,
        "H": 22,
    },
)
for col in range(9, 17):
    ws_call.column_dimensions[get_column_letter(col)].width = 13

add_section_header(ws_call, 5, 1, 5, "Arama Sonuç Dağılımı")
call_result_end_row, _ = write_dataframe_table(
    ws_call,
    call_result_perf,
    6,
    1,
    "CallResultDistributionTable",
    money_columns={"Toplam_Tahsilat_Tutarı"},
    percent_columns={"Yüzde"},
    decimal_columns={"Ortalama_Çağrı_Süresi_Dk"},
    integer_columns={"Arama_Sayısı"},
)

pie = PieChart()
pie_data = Reference(ws_call, min_col=2, min_row=6, max_row=call_result_end_row)
pie_labels = Reference(ws_call, min_col=1, min_row=7, max_row=call_result_end_row)
pie.add_data(pie_data, titles_from_data=True)
pie.set_categories(pie_labels)
style_chart(pie, "Arama Sonuçlarının Yüzdesel Dağılımı", width=15, height=8.5, legend_position="r")
pie.varyColors = True
pie.dLbls = DataLabelList()
pie.dLbls.showPercent = True
pie.dLbls.showLeaderLines = True
pie.series[0].data_points = [
    DataPoint(idx=0, spPr=GraphicalProperties(solidFill=MID_GREY)),
    DataPoint(idx=1, spPr=GraphicalProperties(solidFill=GOLD)),
    DataPoint(idx=2, spPr=GraphicalProperties(solidFill=ANTHRACITE_2)),
    DataPoint(idx=3, spPr=GraphicalProperties(solidFill=EMERALD)),
]
pie.anchor = "G5"
ws_call.add_chart(pie)

debt_columns = [
    "Borç_Tipi",
    "Toplam_Arama",
    "Toplam_Tahsilat_Tutarı",
    "Arama_Başına_Ort_Tahsilat",
    "Ortalama_Çağrı_Süresi_Dk",
    "PTP_Dönüşüm_Oranı",
    "Tahsilat_Başarı_Oranı",
    "Dakika_Başına_Tahsilat",
]
add_section_header(ws_call, 20, 1, 8, "Borç Tipine Göre Tahsilat Başarısı")
debt_end_row, _ = write_dataframe_table(
    ws_call,
    debt_perf[debt_columns],
    21,
    1,
    "DebtTypePerformanceTable",
    money_columns={"Toplam_Tahsilat_Tutarı", "Arama_Başına_Ort_Tahsilat", "Dakika_Başına_Tahsilat"},
    percent_columns={"PTP_Dönüşüm_Oranı", "Tahsilat_Başarı_Oranı"},
    decimal_columns={"Ortalama_Çağrı_Süresi_Dk"},
    integer_columns={"Toplam_Arama"},
)

debt_chart = BarChart()
debt_chart.type = "col"
debt_chart.grouping = "clustered"
debt_chart_data = Reference(ws_call, min_col=3, min_row=21, max_row=debt_end_row)
debt_chart_categories = Reference(ws_call, min_col=1, min_row=22, max_row=debt_end_row)
debt_chart.add_data(debt_chart_data, titles_from_data=True)
debt_chart.set_categories(debt_chart_categories)
style_chart(debt_chart, "Borç Tipine Göre Toplam Tahsilat (₺)", width=14.5, height=8, legend_position="b")
debt_chart.y_axis.numFmt = MONEY0_FMT
debt_chart.y_axis.title = "Tahsilat (₺)"
debt_chart.series[0].graphicalProperties.solidFill = EMERALD
debt_chart.series[0].graphicalProperties.line.solidFill = EMERALD_DARK
debt_chart.anchor = "I20"
ws_call.add_chart(debt_chart)

segment_columns = [
    "Borçlu_Segmenti",
    "Toplam_Arama",
    "Toplam_Tahsilat_Tutarı",
    "Arama_Başına_Ort_Tahsilat",
    "Ortalama_Çağrı_Süresi_Dk",
    "PTP_Dönüşüm_Oranı",
    "Tahsilat_Başarı_Oranı",
    "Dakika_Başına_Tahsilat",
]
add_section_header(ws_call, 36, 1, 8, "Borçlu Segmentine Göre Verimlilik")
segment_end_row, _ = write_dataframe_table(
    ws_call,
    segment_perf[segment_columns],
    37,
    1,
    "DebtorSegmentPerformanceTable",
    money_columns={"Toplam_Tahsilat_Tutarı", "Arama_Başına_Ort_Tahsilat", "Dakika_Başına_Tahsilat"},
    percent_columns={"PTP_Dönüşüm_Oranı", "Tahsilat_Başarı_Oranı"},
    decimal_columns={"Ortalama_Çağrı_Süresi_Dk"},
    integer_columns={"Toplam_Arama"},
)

segment_chart = BarChart()
segment_chart.type = "col"
segment_chart.grouping = "clustered"
segment_chart_data = Reference(ws_call, min_col=3, min_row=37, max_row=segment_end_row)
segment_chart_categories = Reference(ws_call, min_col=1, min_row=38, max_row=segment_end_row)
segment_chart.add_data(segment_chart_data, titles_from_data=True)
segment_chart.set_categories(segment_chart_categories)
style_chart(segment_chart, "Borçlu Segmentine Göre Toplam Tahsilat (₺)", width=14.5, height=8, legend_position="b")
segment_chart.y_axis.numFmt = MONEY0_FMT
segment_chart.y_axis.title = "Tahsilat (₺)"
segment_chart.series[0].graphicalProperties.solidFill = ANTHRACITE_2
segment_chart.series[0].graphicalProperties.line.solidFill = ANTHRACITE
segment_chart.anchor = "I36"
ws_call.add_chart(segment_chart)
ws_call.freeze_panes = "A6"
ws_call.print_area = f"A1:P{max(50, segment_end_row)}"

# ----- Ham Veri --------------------------------------------------------------
style_sheet_base(ws_raw, tab_color=MID_GREY, zoom=82)
add_title(
    ws_raw,
    "HAM VERİ LİSTESİ",
    "Temizlenmiş çağrı kayıtları ve hesaplanmış tahsilat / PTP göstergeleri",
    len(df.columns),
)
raw_end_row, raw_end_col = write_dataframe_table(
    ws_raw,
    df,
    5,
    1,
    "RawCollectionDataTable",
    money_columns={"Alınan_Söz_Tutarı", "Gerçekleşen_Tahsilat", "Açık_Borç_Bakiyesi_Çağrı_Öncesi"},
    decimal_columns={"Çağrı_Süresi_Dk"},
    integer_columns={"PTP_Dönüştü", "Tahsilat_Başarı_Flag"},
    date_columns={"Tarih"},
)
set_widths(
    ws_raw,
    {
        "A": 16,
        "B": 14,
        "C": 13,
        "D": 21,
        "E": 12,
        "F": 18,
        "G": 20,
        "H": 20,
        "I": 24,
        "J": 18,
        "K": 19,
        "L": 30,
        "M": 15,
        "N": 22,
        "O": 13,
    },
)
ws_raw.conditional_formatting.add(
    f"I6:I{raw_end_row}",
    DataBarRule(start_type="num", start_value=0, end_type="max", color=EMERALD, showValue=True),
)
ws_raw.freeze_panes = "D6"
ws_raw.auto_filter.ref = f"A5:{get_column_letter(raw_end_col)}{raw_end_row}"
ws_raw.print_title_rows = "1:5"

# Tüm çalışma sayfalarında tutarlı görünüm ve satır yüksekliği
for ws in wb.worksheets:
    for row in range(1, min(ws.max_row, 100) + 1):
        if ws.row_dimensions[row].height is None:
            ws.row_dimensions[row].height = 19
    ws.sheet_view.view = "normal"

wb.active = wb.sheetnames.index("Özet Dashboard")
wb.save(OUTPUT_FILE)

# Dosya sonrası teknik doğrulama
from openpyxl import load_workbook

check_wb = load_workbook(OUTPUT_FILE, data_only=False, read_only=False)
assert check_wb.sheetnames == ["Özet Dashboard", "Temsilci Performansı", "Çağrı Analizi", "Ham Veri"]
assert check_wb["Ham Veri"].max_row == ROW_COUNT + 5
assert check_wb["Ham Veri"]["A6"].value == "CAG-000001"
assert len(check_wb["Temsilci Performansı"]._charts) == 1
assert len(check_wb["Çağrı Analizi"]._charts) == 3
assert len(check_wb["Özet Dashboard"]._charts) == 1
check_wb.close()

# -----------------------------------------------------------------------------
# 6) Sonuç özeti ve Colab / Jupyter indirme bağlantısı
# -----------------------------------------------------------------------------
print("=" * 78)
print("COLLECTION PERFORMANCE REPORT BAŞARIYLA OLUŞTURULDU")
print("=" * 78)
print(f"Dosya          : {os.path.abspath(OUTPUT_FILE)}")
print(f"Toplam Arama   : {total_calls:,}".replace(",", "."))
print(f"Toplam Tahsilat: {tr_money(total_collection)}")
print(f"Genel PTP      : {tr_percent(overall_ptp_rate)}")
average_call_duration_tr = f"{average_call_duration:.2f}".replace(".", ",")
print(f"Ort. Çağrı     : {average_call_duration_tr} dk")
print("\nYÖNETİCİ ÇIKARIMLARI")
for insight in insights:
    print(insight)

try:
    from google.colab import files as colab_files

    colab_files.download(OUTPUT_FILE)
except ImportError:
    try:
        from IPython.display import FileLink, display

        display(FileLink(OUTPUT_FILE))
    except ImportError:
        print(f"\nExcel dosyası hazır: {os.path.abspath(OUTPUT_FILE)}")
